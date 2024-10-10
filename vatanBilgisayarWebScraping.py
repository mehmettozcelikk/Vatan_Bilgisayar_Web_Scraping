import time
import re
import os
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.common.exceptions import NoSuchElementException
import openpyxl
from openpyxl import Workbook

#TR - System sınıfı, programın ana döngüsünü ve işlevlerini içerir.
#EN - The System class contains the main loop and functions of the program. 
class System:
    def __init__(self):
        self.loop = True #TR - Programın çalışmaya devam edip etmeyeceğini kontrol eder.
        #EN - Controls whether the program continues to run or not.
    
    #TR - Programın ana işleyişi, menüden seçim yapıldıktan sonra başlar.
    #EN - The main operation of the program starts after selecting from the menu.
    def program(self):
        select = self.menu()
        
        #TR - 'q' seçilmediği sürece tarayıcı başlatılır.
        #EN - The browser is initiated unless 'q' is selected.
        if select != "q":
            driver = self.initDriver()

        #TR - Menü fonksiyonundan seçilen seçeneğe göre uygun fonksiyon çağrılır.
        # The appropriate function is called based on the selected option from the menu function.
        if select == "1":
            time.sleep(1)
            self.searchItem(driver)
        if select == "2":
            time.sleep(1)
            self.phoneTab(driver)
        if select == "3":
            time.sleep(1)
            self.compTab(driver)
        if select == "4":
            time.sleep(1)
            self.tvTab(driver)
        if select == "5":
            time.sleep(1)
            self.compPartsTab(driver)
        if select == "6":
            time.sleep(1)
            self.houseItemsTab(driver)
        if select == "7":
            time.sleep(1)
            self.personalCareTab(driver)        
        if select == "q":
            time.sleep(1)
            self.quit()

    #TR - Kullanıcıdan menü seçimini alır ve girdiyi doğrular.
    #EN - Prompts the user for menu selection and validates the input.
    def menu(self):
        #TR - Geçersiz girişler için hata fırlatılır.
        #EN - Throws an error for invalid inputs.
        def controlSelect(select):
            if re.search("[^1-7]",select) and select != "q":
                raise Exception("Lütfen geçerli bir seçim yapınız!!")
            if len(select) != 1:
                raise Exception("Lütfen geçerli bir seçim yapınız!!")

        print("\nMerhaba, Vatan Bilgisayar veri kazıma programına hoşgeldiniz.\n")
        
        while True:
            try:
                select = input("[1]-Ürün Araması Yap\n[2]-Cep Telefonlarını Göster\n[3]-Bilgisayarları Göster\n[4]-Televizyonları Göster\n[5]-Bilgisayar Parçalarını Göster\n[6]-Ev&Mutfak Ürünlerini Göster\n[7]-Kişisel Bakım Ürünlerini Göster\n[q]-Çıkış Yap\nLütfen yapmak istediğiniz işlemi seçiniz:")
                controlSelect(select)
            except Exception as error:
                print(error)
                time.sleep(2)
            else:
                break
        return select
    
    #TR - Ana menüye dönme veya çıkış işlemi sağlar.
    #EN - Provides functionality to return to the main menu or exit.
    def returnMenu(self):
        def controlSelect(select):
            if select != 1 and select != 2:
                raise Exception("Hatalı karakter girişi!!")
        
        while True:
            try:
                select = int(input("Ana menüye dönmek için 1'i,çıkış yapmak için 2'yi tuşlayınız:"))
                controlSelect(select)
                
                if select == 1:
                    print("Ana menüye dönülüyor..")
                    time.sleep(2)
                    self.program()
                if select == 2:
                    self.quit()
            except Exception as error:
                print(error)
                time.sleep(2)
            else:
                break
    
    #TR - Programı kapatır.
    #EN - Exits the program.
    def quit(self):
        print("Çıkış yapılıyor..\nİyi günler :)")
        self.loop = False
        exit()
        time.sleep(2)

    #TR - Selenium ile tarayıcıyı başlatır ve siteyi açar.
    #EN - Initiates the browser using Selenium and opens the website.
    def initDriver(self):
        service = Service(executable_path="C:/Users/Mehmet Özçelik/Desktop/chromedriver.exe")
        driver = webdriver.Chrome(service=service)
        url = "https://www.vatanbilgisayar.com/"
        driver.get(url)
        print("Siteye bağlanıyor..\n")
        time.sleep(2)
        driver.maximize_window()
        return driver
    
    #TR - Web sayfasındaki ürünleri bulur ve kaydeder.
    #EN - Finds and saves products from the webpage.
    def findElement(self,driver,sheet_name):

        pageNum = 1
        #TR - Mevcut URL'yi kontrol ederek sayfalama yapısını belirler ve bir sonraki sayfayı yükler.
        #EN - Checks the current URL to determine pagination structure and loads the next page.
        while True:
            currentUrl = driver.current_url
            if "?page=" in currentUrl:
                pageUrl = currentUrl.split("?page=")[0] + f"?page={pageNum}"
            else:
                pageUrl = currentUrl + f"?page={pageNum}"
            driver.get(pageUrl)
            pageNum += 1
            time.sleep(2)
        
            try:
                #TR - Ürün listesini seçer. (Tüm ürünleri kapsayan div'i alır.)
                #EN - Selects the product list. (Gets the div that contains all products.)
                products = driver.find_elements(By.CSS_SELECTOR,".product-list.product-list--list-page")

                #TR - Ürün yoksa döngüyü sonlandırır.
                #EN - If there are no more products, break the loop.
                if not products:
                    print("Daha fazla ürün bulunamadı!")
                    break
                
                index = 0
                productsLen = len(products)
                productNames = []
                productPrices = []

                for i in products:
                    if index != productsLen:
                        index += 1
                    
                    #TR - Ürün ismi bulmaya çalışır, bulunamazsa devam eder.
                    #EN - Tries to find product names, continues if not found.
                    try:
                        pName = i.find_element(By.XPATH,"//*[@id='productsLoad']/div[{}]/div/div[3]/h3".format(index)).text
                    except NoSuchElementException:
                        print("Ürün ismi bulunamadı.")
                        continue
                    productNames.append(pName)

                    #TR - Ürün fiyatını bulmaya çalışır, farklı xpath yollarını dener. (web sayfasının yapısından kaynaklı)
                    #EN - Tries to find product prices, tries different xpath paths.(due to the structure of the web page)
                    try:
                        pPrice = i.find_element(By.XPATH,"//*[@id='productsLoad']/div[{}]/div/div[4]/div[1]/span[1]".format(index)).text 
                    except NoSuchElementException:
                        try:
                            pPrice = i.find_element(By.XPATH,"//*[@id='productsLoad']/div[{}]/div[1]/div[4]/div[1]/span[1]".format(index)).text
                        except NoSuchElementException:
                            try:
                                pPrice = i.find_element(By.XPATH,"//*[@id='productsLoad']/div[{}]/div/div[4]/div[1]/div/div[2]/span[1]".format(index)).text
                            except NoSuchElementException:
                                try:
                                    pPrice = i.find_element(By.XPATH,"//*[@id='productsLoad']/div[{}]/div/div[5]/div[1]/span[1]".format(index)).text
                                except NoSuchElementException:
                                    print("Ürün fiyatı bulunamadı.")
                                    continue
                    productPrices.append(pPrice)
                products.clear()
                time.sleep(2)
            except NoSuchElementException:
                break
            #TR - Excel dosyasına ürünleri kaydeder.
            #EN - Saves products to the Excel file.
            self.saveExcel(sheet_name,productNames,productPrices)
            productNames.clear()
            productPrices.clear()
        driver.close()
        self.returnMenu()

    #TR - Arama kutusuna kullanıcı tarafından girilen anahtar kelime ile arama yapar.
    #EN - Searches for a user-entered keyword in the search bar.
    def searchItem(self,driver):
        key = input("\nAramak istediğiniz ürünü giriniz:")

        searchButton = driver.find_element(By.XPATH,"//*[@id='navbar-search-input']")
        searchButton.send_keys(key)
        time.sleep(1)
        searchButton.send_keys(Keys.ENTER)
        time.sleep(2)
        
        print(f"{key} ile ilgili veriler alınıyor. Lütfen bekleyiniz...")
        
        #TR - Ürünleri bulur ve excel dosyasına kaydeder.
        #EN - Finds products and saves them in an excel file.
        self.findElement(driver,key)

    # TR - Telefon kategorisini açar ve ürün bilgilerini toplar.
    # EN - Opens the phone category and collects product information.
    def phoneTab(self,driver):
        phoneButton = driver.find_element(By.XPATH,"//*[@id='navbar']/ul/li[1]/a")
        phoneButton.click()
        time.sleep(2)
    
        sheet_name = driver.find_element(By.XPATH,"//*[@id='product-list-container']/div/div/div[2]/div[1]/h1").text

        print("Telefon kategorisindeki veriler alınıyor. Lütfen bekleyiniz...")

        self.findElement(driver,sheet_name)

    # TR - Bilgisayar kategorisini açar ve ürün bilgilerini toplar.
    # EN - Opens the computer category and collects product information.
    def compTab(self,driver):
        compButton = driver.find_element(By.XPATH,"//*[@id='navbar']/ul/li[2]/a")
        compButton.click()
        time.sleep(2)

        sheet_name = driver.find_element(By.XPATH,"//*[@id='product-list-container']/div/div/div[2]/div[1]/h1").text

        print("Bilgisayar kategorisindeki veriler alınıyor. Lütfen bekleyiniz...")

        self.findElement(driver,sheet_name)

    # TR - Televizyon kategorisini açar ve ürün bilgilerini toplar.
    # EN - Opens the television category and collects product information.
    def tvTab(self,driver):
        tvButton = driver.find_element(By.XPATH,"//*[@id='navbar']/ul/li[3]/a")
        tvButton.click()
        time.sleep(2)

        sheet_name = driver.find_element(By.XPATH,"//*[@id='product-list-container']/div/div/div[2]/div[1]/h1").text

        print("Televizyon kategorisindeki veriler alınıyor. Lütfen bekleyiniz...")

        self.findElement(driver,sheet_name)

    # TR - Bilgisayar parçaları kategorisini açar ve ürün bilgilerini toplar.
    # EN - Opens the computer parts category and collects product information.
    def compPartsTab(self,driver):
        tvButton = driver.find_element(By.XPATH,"//*[@id='navbar']/ul/li[4]/a")
        tvButton.click()
        time.sleep(2)

        sheet_name = driver.find_element(By.XPATH,"//*[@id='product-list-container']/div/div/div[2]/div[1]/h1").text

        print("Bilgisayar parçaları kategorisindeki veriler alınıyor. Lütfen bekleyiniz...")

        self.findElement(driver,sheet_name)

    # TR - Ev ve mutfak ürünleri kategorisini açar, kullanıcı seçimine göre alt kategorilere gider ve ürün bilgilerini toplar.
    # EN - Opens the house and kitchen items category, navigates to subcategories based on user selection, and collects product information.
    def houseItemsTab(self,driver):
        houseButton = driver.find_element(By.XPATH,"//*[@id='navbar']/ul/li[5]/a")
        houseButton.click()
        time.sleep(1)

        # TR - Çerezleri kabul eder.
        # EN - Accepts cookies.
        acceptCookies = driver.find_element(By.XPATH,"//*[@id='ccp---nb']/div[1]/div[2]/button[1]")
        acceptCookies.click()
        time.sleep(1)
        
        # # TR - Kullanıcıdan gelen alt kategori seçimini kontrol eder. (Mutfak Ürünleri veya Elektrikli Ev Aletleri).
        # EN - Controls the subcategory selection from the user. (Kitchen Products or Electrical Appliances).
        def controlSelectTab(selectTab):
            if selectTab != 1 and selectTab != 2:
                raise Exception("Hatalı karakter girişi!!")
        
        # TR - Mutfak ürünleri kategorisine tıklayıp seçim yapar ve veri toplar.
        # EN - Clicks the kitchen products category, makes a selection, and collects data.
        def kitchenTab():
            def controlKitchenSelect(kitchenSelect):
                if not str(kitchenSelect).isdigit() or not 0 < kitchenSelect <= 18:
                    raise Exception("Hatalı karakter girişi!!")
            
            kitchenButton = driver.find_element(By.XPATH,"//*[@id='v-pills-mutfak-tab']")
            kitchenButton.click()
            time.sleep(1)

            while True:
                try:
                    kitchenSelect = int(input("[1]-Airfyer\n[2]-İçecek Ürünleri\n[3]-Blender\n[4]-Meyve Sıkacakları\n[5]-Türk Kahvesi Makineleri\n[6]-Ekmek Kızartma\n[7]-Tost Makineleri\n[8]-Su Isıtıcı ve Kettle\n[9]-Kıyma Makineleri\n[10]-Mutfak Robotları\n[11]-Izgara Makineleri\n[12]-Waffle Makineleri\n[13]-Mutfak Tartıları\n[14]-Rondolar ve Doğrayıcılar\n[15]-Yumurta Makineleri\n[16]-Fırınlar\n[17]-Mikser\n[18]-Mutfak Gereçleri\nVerilerini almak istediğiniz sekmeyi tuşlayınız:"))
                    controlKitchenSelect(kitchenSelect)
                    
                    choose = driver.find_element(By.XPATH,"//*[@id='Mutfak']/div/div[{}]/div[2]/a".format(kitchenSelect))
                    choose.click()
                    time.sleep(1)
                            
                    sheet_name = driver.find_element(By.XPATH,"//*[@id='product-list-container']/div/div/div[2]/div[1]/h1").text
                    print(f"{sheet_name} bilgileri alınıyor..")
                    self.findElement(driver,sheet_name)
                except Exception as error:
                    print(error)
                    time.sleep(2)
                else:
                    break
        
        # TR - Elektrikli ev aletleri kategorisine tıklayıp seçim yapar ve veri toplar.
        # EN - Clicks the electrical appliances category, makes a selection, and collects data.
        def electricalAppliances():
            def controlElectricalSelect(electricalSelect):
                if not str(electricalSelect).isdigit() or not 1<= electricalSelect <= 18:
                    raise Exception("Hatalı karakter girişi!!")
            
            electricalButton = driver.find_element(By.XPATH,"//*[@id='v-pills-elektrikli-tab']")
            electricalButton.click()
            time.sleep(1)

            while True:
                try:
                    electricalSelect = int(input("[1]-Robot Süpürgeler\n[2]-Dikey Süpürgeler\n[3]-Toz Torbalı Süpürgeler\n[4]-Toz Torbasız Süpürgeler\n[5]-Şarjlı Süpürgeler\n[6]-Buharlı Ütüler\n[7]-Dikey Ütüler\n[8]-Buhar Kazanlı Ütüler\n[9]-Kıyma Makineleri\n[10]-Buharlı Temizleyciler\n[11]-Hava Temizleyici\n[12]-Ütü Masası\n[13]-Isıtıcılar\n[14]-Su Sebili\n[15]-Vantilatör\n[16]-Şofben\n[17]-Termosifon\n[18]-Aksesuarlar\nVerilerini almak istediğiniz sekmeyi tuşlayınız:"))
                    controlElectricalSelect(electricalSelect)

                    choose = driver.find_element(By.XPATH,"//*[@id='ElektrikliEv']/div/div[{}]/div[2]/a".format(electricalSelect))
                    choose.click()
                    time.sleep(2)

                    sheet_name = driver.find_element(By.XPATH,"//*[@id='product-list-container']/div/div/div[2]/div[1]/h1").text
                    print(f"{sheet_name} bilgileri alınıyor..")
                    self.findElement(driver,sheet_name)
                except Exception as error:
                    print(error)
                    time.sleep(2)
                else:
                    break

        while True:
            try:
                selectTab = int(input("[1]-Mutfak Ürünleri\n[2]-Elektrikli Ev Aletleri\nLütfen verisine ulaşmak istediğiniz sekmeyi seçiniz:"))
                controlSelectTab(selectTab)
                try:
                    if selectTab == 1:
                        print("Mutfak ürünleri sekmesine gidiliyor..")
                        time.sleep(2)  
                        kitchenTab()
                    if selectTab == 2:
                        print("Elektrikli ev aletleri sekmesine gidiliyor..")
                        electricalAppliances()
                except NoSuchElementException:
                    print("Sekme bulunamadı!")
                    break 
            except Exception as error:
                print(error)
                time.sleep(2)
            else:
                break
    
    # TR - Kişisel bakım kategorisini açar ve ürün bilgilerini toplar.
    # EN - Opens the personal care category and collects product information.
    def personalCareTab(self,driver):
        pCareButton = driver.find_element(By.XPATH,"//*[@id='navbar']/ul/li[6]/a")
        pCareButton.click()
        time.sleep(1)

        sheet_name = driver.find_element(By.XPATH,"//*[@id='product-list-container']/div/div/div[2]/div[1]/h1").text

        print("Kişisel Bakım kategorisindeki veriler alınıyor. Lütfen bekleyiniz...")

        self.findElement(driver,sheet_name)

    #TR - Alınan verileri excel dosyasına kaydeder.
    #EN - Saves the received data to an excel file.
    def saveExcel(self,sheetName,productNames,productPrices):
        excelPath = "C:/Users/Mehmet Özçelik/Desktop/vatanData.xlsx"
        #TR - Eğer Excel dosyası mevcutsa, workbook'u aç
        #EN - If the Excel file exists, open the workbook 
        if os.path.exists(excelPath):
            wb = openpyxl.load_workbook(excelPath)
            #TR - Eğer belirtilen sayfa varsa, sayfayı aç, yoksa yeni sayfa oluştur
            #EN - If the specified sheet exists, open it, otherwise create a new sheet
            if sheetName in wb.sheetnames:
                ws = wb[sheetName]
            else:
                ws = wb.create_sheet(title=sheetName)
        else: #TR - Eğer dosya mevcut değilse, yeni workbook oluştur ve sayfa ismini ayarla
        #EN - If the file does not exist, create a new workbook and set the sheet name
            wb = Workbook()
            ws = wb.active
            ws.title = sheetName
        
        #TR - Mevcut satır sayısını alarak, yeni verileri o satırdan başlatır.
        #EN - Get the current number of rows and start adding new data from that row.
        existingRows = ws.max_row

        #TR - Ürün isimlerini belirtilen satırlara ekle
        #EN - Add product names to the specified rows
        for index,item in enumerate(productNames,start=existingRows+1):
            ws.cell(row=index,column=1,value=item)

        #TR - Ürün fiyatlarını belirtilen satırlara ekle
        #EN - Add product prices to the specified rows
        for index,item in enumerate(productPrices,start=existingRows+1):
            ws.cell(row=index,column=2,value=item)
        
        wb.save(excelPath)
        wb.close()

system = System()

#TR - Programı çalıştır.
#EN - Run the program.
while True:
    system.program()
